#!/usr/bin/env python3
"""
This script runs the GEDLIB executable and captures its output line-by-line,
parses that output, and writes the results to an Excel file as soon as possible.
Intermediate results are flushed frequently so that if the process is terminated
(e.g. via Ctrl+C), the work done so far is saved. The Excel file is written using
a fallback strategy and will be split into multiple files if the data exceeds Excel’s
maximum row capacity.
"""

import os
import re
import subprocess
import tempfile
import xml.etree.ElementTree as ET
import platform
import pandas as pd
import signal
import sys
import psutil  # For polling subprocess memory usage
from typing import Optional

if platform.system() != "Windows":
    import resource

# --------------------------
# Global variables and settings
# --------------------------
global_results = []            # Intermediate results list.
global_ged_process: Optional[subprocess.Popen] = None  # Handle to the GEDLIB subprocess.
global_preprocessed_xml: Optional[str] = None          # Path to the temporary preprocessed XML.

# Modify these paths as needed:
GED_EXECUTABLE = "/home/mfilippov/CLionProjects/gedlib/build/main_exec"
DATASET_PATH = "/home/mfilippov/ged_data/processed_data/gxl/IMDB-BINARY"
COLLECTION_XML = "/home/mfilippov/ged_data/processed_data/xml/IMDB-BINARY.xml"
RESULTS_DIR = "/home/mfilippov/ged_data/results/gedlib/IMDB-BINARY/IPFP"
RESULTS_FILE = os.path.join(RESULTS_DIR, "IMDB-BINARY_IPFP_results_1.xlsx")
EXACT_GED_FILE = "/home/mfilippov/ged_data/results/exact_ged/IMDB-BINARY/results.xlsx"

# Mapping of method ID to method names.
METHOD_NAMES = {
    8:  "Anchor Aware",
    10: "IPFP",
    11: "BIPARTITE",
    16: "REFINE",
    19: "HED",
    20: "STAR (Exact)"
}
# Maximum number of rows per Excel file.
EXCEL_MAX_ROWS = 1048573

# Number of graph pairs to skip.
SKIP_PAIRS = 0

# --------------------------
# Utility Functions
# --------------------------
def set_unlimited():
    """Set resource limits to unlimited (if supported)."""
    if platform.system() != "Windows":
        try:
            resource.setrlimit(resource.RLIMIT_AS, (resource.RLIM_INFINITY, resource.RLIM_INFINITY))
        except Exception as e:
            print("Warning: could not set RLIMIT_AS unlimited:", e)
        try:
            resource.setrlimit(resource.RLIMIT_CPU, (resource.RLIM_INFINITY, resource.RLIM_INFINITY))
        except Exception as e:
            print("Warning: could not set RLIMIT_CPU unlimited:", e)

def save_results(excel_file, results_list):
    """
    Save the results list to an Excel file.
    This function attempts to save using openpyxl first and falls back to xlsxwriter.
    If both fail, it falls back to saving as CSV.
    Splitting into multiple files is performed if the DataFrame exceeds Excel's row limit.
    Any errors (such as "I/O operation on closed file") are caught and handled.
    """
    try:
        df = pd.DataFrame(results_list)
        # Standardize column names.
        df["graph_id_1"] = df["graph1"]
        df["graph_id_2"] = df["graph2"]
        desired_columns = [
            "method", "graph_id_1", "graph_id_2", "ged", "accuracy",
            "absolute_error", "squared_error", "runtime", "memory_usage_mb",
            "graph1_n", "graph1_density", "graph2_n", "graph2_density", "scalability"
        ]
        df = df[desired_columns]
        os.makedirs(os.path.dirname(excel_file), exist_ok=True)
    except Exception as e:
        print("Error building DataFrame:", e)
        return

    def attempt_save(engine, file_path):
        temp_file = os.path.join(os.path.dirname(file_path), "temp_results.xlsx")
        try:
            df.to_excel(temp_file, index=False, engine=engine)
            os.replace(temp_file, file_path)
            print(f"Results saved to {file_path} using {engine}.")
            return True
        except Exception as ex:
            print(f"Error saving with {engine}: {ex}")
            if os.path.exists(temp_file):
                os.remove(temp_file)
            return False

    # Split into multiple files if needed.
    if len(df) > EXCEL_MAX_ROWS:
        num_files = (len(df) + EXCEL_MAX_ROWS - 1) // EXCEL_MAX_ROWS
        print(f"DataFrame has {len(df)} rows; splitting into {num_files} files.")
        for part in range(num_files):
            start = part * EXCEL_MAX_ROWS
            end = start + EXCEL_MAX_ROWS
            chunk = df.iloc[start:end]
            if part == 0:
                file_path = excel_file
            else:
                base, ext = os.path.splitext(excel_file)
                file_path = f"{base}_part{part+1}{ext}"
            if os.path.exists(file_path):
                try:
                    from openpyxl import load_workbook
                    load_workbook(file_path)
                except Exception as ex:
                    print(f"Existing file {file_path} is corrupted ({ex}). Removing it.")
                    os.remove(file_path)
            if not attempt_save("openpyxl", file_path):
                if not attempt_save("xlsxwriter", file_path):
                    print(f"Failed to save {file_path} using both engines.")
        return
    else:
        file_path = excel_file
        if os.path.exists(file_path):
            try:
                from openpyxl import load_workbook
                load_workbook(file_path)
            except Exception as e:
                print(f"Existing file {file_path} is corrupted ({e}). Removing it.")
                os.remove(file_path)
        if not attempt_save("openpyxl", file_path):
            if not attempt_save("xlsxwriter", file_path):
                print("Failed to save Excel file using both engines. Attempting to save as CSV.")
                try:
                    csv_file = os.path.splitext(file_path)[0] + ".csv"
                    df.to_csv(csv_file, index=False)
                    print(f"Results saved to {csv_file} as CSV fallback.")
                except Exception as ex:
                    print(f"Failed to save results as CSV: {ex}")

def signal_handler(signum, frame):
    """Handle termination signals by cleaning up and saving partial results."""
    print(f"\nSignal {signum} received. Saving intermediate results and exiting.")
    global global_ged_process, global_preprocessed_xml
    if global_ged_process is not None:
        try:
            global_ged_process.terminate()
            global_ged_process.wait(timeout=5)
        except Exception as e:
            print("Error terminating GED subprocess:", e)
    if global_preprocessed_xml is not None and os.path.exists(global_preprocessed_xml):
        try:
            os.remove(global_preprocessed_xml)
        except Exception as e:
            print("Error removing temporary XML file:", e)
    save_results(RESULTS_FILE, global_results)
    sys.exit(1)

# Register signal handlers for graceful termination.
signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)
if hasattr(signal, "SIGHUP"):
    signal.signal(signal.SIGHUP, signal_handler)

def preprocess_xml_file(xml_path: str) -> str:
    """
    Remove any DOCTYPE declarations from the XML file to avoid parsing issues,
    and return the path to a temporary file with the cleaned XML.
    """
    try:
        with open(xml_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
    except Exception as e:
        print(f"Error reading XML file '{xml_path}': {e}")
        raise
    filtered_lines = [line for line in lines if not line.lstrip().startswith("<!DOCTYPE")]
    temp_fd, temp_path = tempfile.mkstemp(suffix=".xml")
    with os.fdopen(temp_fd, "w", encoding="utf-8") as f:
        f.writelines(filtered_lines)
    return temp_path

def get_graph_properties(gxl_file: str):
    """
    Parse the given GXL file and compute:
      - number of nodes (n)
      - number of edges (e)
      - density = 2*e / (n*(n-1)) for undirected graphs (if n > 1)
    """
    try:
        temp_path = preprocess_xml_file(gxl_file)
        tree = ET.parse(temp_path)
        os.remove(temp_path)
    except Exception as e:
        print(f"Error parsing {gxl_file}: {e}")
        return None, None, None

    root = tree.getroot()
    graph_elem = root.find('graph')
    if graph_elem is None:
        print(f"No <graph> element found in {gxl_file}.")
        return None, None, None
    nodes = graph_elem.findall('node')
    edges = graph_elem.findall('edge')
    n = len(nodes)
    e = len(edges)
    density = (2 * e) / (n * (n - 1)) if n > 1 else 0
    return n, e, density

def get_first_two_graph_properties(dataset_path: str, collection_xml: str):
    """
    Parse the collection XML (after cleaning) to retrieve the first two graph file names,
    then compute and return their (n, edges, density) properties.
    """
    try:
        temp_xml = preprocess_xml_file(collection_xml)
        tree = ET.parse(temp_xml)
        os.remove(temp_xml)
    except Exception as e:
        print(f"Error parsing collection XML: {e}")
        return None, None

    root = tree.getroot()
    graphs = root.findall('graph')
    if len(graphs) < 2:
        print("Not enough graphs found in collection XML.")
        return None, None
    file1 = graphs[0].get('file')
    file2 = graphs[1].get('file')
    if not file1 or not file2:
        print("Graph file names missing in collection XML.")
        return None, None
    path1 = os.path.join(dataset_path, file1)
    path2 = os.path.join(dataset_path, file2)
    props1 = get_graph_properties(path1)
    props2 = get_graph_properties(path2)
    return props1, props2

def run_ged(dataset_path: str, collection_xml: str):
    """
    Run the GEDLIB executable with the given dataset and collection XML.
    Parse its output line-by-line and flush intermediate results every few pairs.
    On termination or error, results are saved and temporary files are cleaned up.
    """
    global global_ged_process, global_preprocessed_xml
    try:
        preprocessed_xml = preprocess_xml_file(collection_xml)
        global_preprocessed_xml = preprocessed_xml
    except Exception as e:
        print("Failed to preprocess collection XML:", e)
        return [{"error": "Preprocessing XML failed"}]

    command = [GED_EXECUTABLE, dataset_path, preprocessed_xml]
    print("Running command:", " ".join(command))
    try:
        process = subprocess.Popen(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            bufsize=1,
            preexec_fn=set_unlimited if platform.system() != "Windows" else None
        )
        global_ged_process = process
    except Exception as e:
        print("Error starting GED subprocess:", e)
        os.remove(preprocessed_xml)
        global_preprocessed_xml = None
        return [{"error": str(e)}]

    try:
        ged_proc = psutil.Process(process.pid)
    except Exception as e:
        print("Error creating psutil.Process for GEDLIB:", e)
        ged_proc = None

    props = get_first_two_graph_properties(dataset_path, collection_xml)
    if props is not None:
        (n1, _, d1), (n2, _, d2) = props
        scalability = max(n1, n2)
    else:
        scalability = None
        n1 = n2 = d1 = d2 = None

    # Define a regex pattern for a floating point number or "inf"
    float_re = r"([+-]?(?:\d+(?:\.\d+)?|\.\d+)(?:[eE][+-]?\d+)?|inf)"
    # Updated regex pattern to match the executable's output line
    regex = re.compile(
        rf"METHOD=(\d+)\s+GRAPH1=(\d+)\s+GRAPH2=(\d+)\s+PREDGED={float_re}\s+GTGED=N/A\s+RUNTIME={float_re}(?:\s+MEM=\S+)?"
    )

    line_count = 0       # Total lines read.
    processed_count = 0  # Count of graph pairs processed (after skipping).
    flush_interval = 5   # Flush intermediate results every 5 processed pairs.
    try:
        for line in process.stdout:
            line = line.strip()
            if not line:
                continue
            line_count += 1
            # Skip the first SKIP_PAIRS graph pairs.
            if line_count <= SKIP_PAIRS:
                continue
            match = regex.search(line)
            if match:
                processed_count += 1
                method_id = int(match.group(1))
                graph1 = int(match.group(2))
                graph2 = int(match.group(3))
                pred_ged = float(match.group(4))
                runtime = float(match.group(5))
                try:
                    # Calculate memory usage internally using psutil.
                    memory_usage_mb = ged_proc.memory_info().rss / (1024 * 1024) if ged_proc else "N/A"
                except Exception:
                    memory_usage_mb = "N/A"
                method_name = METHOD_NAMES.get(method_id, f"Unknown Method {method_id}")
                result_entry = {
                    "method": method_name,
                    "graph1": graph1,
                    "graph2": graph2,
                    "ged": pred_ged,
                    "accuracy": "N/A",         # Optionally compute later using exact GED lookup.
                    "absolute_error": "N/A",
                    "squared_error": "N/A",
                    "runtime": runtime,
                    "memory_usage_mb": memory_usage_mb,
                    "graph1_n": n1 if n1 is not None else "N/A",
                    "graph1_density": round(d1, 4) if d1 is not None else "N/A",
                    "graph2_n": n2 if n2 is not None else "N/A",
                    "graph2_density": round(d2, 4) if d2 is not None else "N/A",
                    "scalability": scalability if scalability is not None else "N/A"
                }
                global_results.append(result_entry)
            else:
                print("Warning: Unmatched line:", line)

            if processed_count % flush_interval == 0 and processed_count != 0:
                save_results(RESULTS_FILE, global_results)
    except Exception as e:
        print("Error while processing GED output:", e)
    finally:
        try:
            process.stdout.close()
        except Exception:
            pass
        process.wait()
        try:
            stderr = process.stderr.read()
            if stderr:
                print("GEDLIB stderr:", stderr)
        except Exception:
            pass
        try:
            process.stderr.close()
        except Exception:
            pass
        save_results(RESULTS_FILE, global_results)
        if global_preprocessed_xml is not None and os.path.exists(global_preprocessed_xml):
            try:
                os.remove(global_preprocessed_xml)
                global_preprocessed_xml = None
            except Exception as e:
                print("Error removing temporary XML file:", e)
    return global_results

if __name__ == "__main__":
    try:
        results = run_ged(DATASET_PATH, COLLECTION_XML)
        if results:
            print(f"Finished processing {len(results)} result(s).")
        else:
            print("No results processed.")
    except Exception as e:
        print("Unexpected error:", e)
        save_results(RESULTS_FILE, global_results)
        sys.exit(1)
